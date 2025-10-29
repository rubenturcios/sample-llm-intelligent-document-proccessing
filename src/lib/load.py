from enum import StrEnum, auto
import logging
from itertools import zip_longest
import os
from pprint import pformat
import tempfile
from typing import Optional
from io import BytesIO

import boto3
import openpyxl
from langchain_core.documents.base import Document
from langchain_community.document_loaders.blob_loaders import Blob
from langchain_community.document_loaders.pdf import (
    PDFMinerPDFasHTMLLoader,
    AmazonTextractPDFLoader
)
from langchain_community.document_loaders.word_document import (
    Docx2txtLoader,
    UnstructuredWordDocumentLoader
)
from langchain_community.document_loaders.parsers.pdf import (
    PyPDFParser,
    PDFMinerParser
)

from .constants import VECTOR_FIELD_NAME
from .utils import load_s3_file
from .opensearch import get_hits_by_msearch, put_os_entries, get_os_entries


format = '%(levelname)s:%(filename)s:%(lineno)d:%(asctime)s:%(message)s'
logging.basicConfig(format=format)
logger = logging.getLogger(__name__)
logger.setLevel(level=os.getenv('LOG_LEVEL', logging.INFO))

session = boto3.Session()
credentials = session.get_credentials()
s3_client = session.client('s3')


class Method(StrEnum):
    PDF_MINER = auto()
    PYPDF = auto()
    AWS_TEXTRACT = auto()
    UNSTRUCTUED = auto()
    DOCX2TXT = auto()


def pdfminer_to_html_load(
    *,
    bucket: str,
    filename: str | None,
    stream: str | bytes | None = None
) -> list[Document]:
    with tempfile.NamedTemporaryFile() as temp_file:
        if stream:
            temp_file.write(stream)
        else:
            data = load_s3_file(bucket, filename).read()
            temp_file.write(data)

        loader = PDFMinerPDFasHTMLLoader(temp_file.name)
        docs = loader.load()

    return docs


def pypdf_load(    
    bucket: str,
    filename: str | None,
    *,
    stream: str | bytes | None = None
) -> list[Document]:
    if stream:
        blob = Blob(data=stream)
    else:
        data = load_s3_file(bucket, filename)
        blob = Blob(data=data.read())

    parser = PyPDFParser()
    docs = parser.parse(blob)
    logger.info('Num of Docs: %s', len(docs))

    for doc in docs:
        doc.metadata['source'] = filename
        doc.metadata['method'] = Method.AWS_TEXTRACT
        logger.debug(pformat(doc.dict()))

    return docs


def pdfminer_load(
    *,
    bucket: str | None = None,
    filename: str | None = None,
    stream: str | bytes | None = None
) -> list[Document]:
    if stream:
        blob = Blob(data=stream)
    else:
        data = load_s3_file(bucket, filename)
        blob = Blob(data=data.read())

    parser = PDFMinerParser(concatenate_pages=False)
    docs = parser.parse(blob)
    logger.info('Num of Docs: %s', len(docs))

    for doc in docs:
        doc.metadata['source'] = filename
        doc.metadata['method'] = str(Method.PDF_MINER)
        logger.debug(pformat(doc.dict()))

    return docs


def awstextract_load(bucket: str, filename: str) -> list[Document]:
    s3_path = f's3://{bucket}/{filename}'

    loader = AmazonTextractPDFLoader(
        file_path=s3_path,
        credentials_profile_name=session.profile_name,
        region_name=session.region_name
    )

    docs = loader.load()
    logger.info('Num of Docs: %s', len(docs))

    for doc in docs:
        doc.metadata['source'] = filename
        doc.metadata['method'] = str(Method.AWS_TEXTRACT)
        logger.debug(pformat(doc.dict()))

    return docs


def unstructructed_load(
    *,
    bucket: str | None = None,
    filename: str | None = None,
    stream: str | bytes | None = None
) -> list[Document]:
    with tempfile.NamedTemporaryFile() as temp_file:
        if stream:
            temp_file.write(stream)
        else:
            data = load_s3_file(bucket, filename).read()
            temp_file.write(data)

        loader = UnstructuredWordDocumentLoader(temp_file.name, 'paged')
        docs = loader.load()

        for doc in docs:
            doc.metadata['source'] = filename
            doc.metadata['method'] = str(Method.UNSTRUCTUED)
            logger.debug(pformat(doc.dict()))

        logger.info('Num of Docs: %s', len(docs))
        return docs


def docx2txt_load(bucket: str, filename:str) -> list[Document]:
    with tempfile.NamedTemporaryFile() as temp_file:
        data = load_s3_file(bucket, filename).read()
        temp_file.write(data)
        loader = Docx2txtLoader(temp_file.name)
        docs = loader.load()

        for doc in docs:
            doc.metadata['source'] = filename
            doc.metadata['method'] = str(Method.DOCX2TXT)
            logger.debug(pformat(doc.dict()))

        logger.info('Num of Docs: %s', len(docs))
        return docs


def upload_to_os(docs: list[Document], index_name: str) -> None:
    entries = [doc.dict() for doc in docs]

    for entry in entries:
        logger.debug(entry)

    return put_os_entries(index_name, entries)


def get_spec_sheet_docs(
    index_name: str,
    file_name: str,
    *,
    method: Method = Method.PDF_MINER,
    size: int = 10000
) -> list[Document]:
    query = {
        "query": {
            "bool": {
                "filter": [
                    {"term": {"metadata.source.keyword": file_name}},
                    {"term": {"metadata.method.keyword": str(method)}}
                ]
            }
        },
        "sort": {
            "metadata.page.keyword": "asc"
        },
        "size": size
    }
    hits = get_os_entries(index_name, query)
    docs = [Document(**doc_raw['_source']) for doc_raw in hits]
    logger.info('Number of %s docs retrieved: %s', file_name, len(docs))

    for doc in docs:
        logger.debug(doc)

    return docs


def get_bid_review_docs(
    index_name: str,
    file_name: str,
    *,
    size: int = 10000
) -> list[Document]:
    query = {
        "query": {
            "term": {"spec_sheet.keyword": file_name}
        },
        "sort": {
            "PAGE.keyword": "asc"
        },
        "size": size
    } 
    hits = get_os_entries(index_name, query)
    docs = [
        Document(
            page_content=doc_raw['_source']['EXCEPTIONS / SPECIAL REQUIREMENTS'],
            metadata=doc_raw['_source'],
        ) for doc_raw in hits
    ]
    logger.info('Number of docs retrieved: %s', len(docs))

    for doc in docs:
        logger.debug(doc)

    return docs


def open_wb(workbook_name: BytesIO, sheet_name: Optional[str] = None) -> None:
    wb = openpyxl.load_workbook(workbook_name)
    sheet = wb[sheet_name] if sheet_name else wb.active
    sheet.delete_cols(0)
    sheet.delete_rows(1)

    rows = []
    column_names = tuple(column.value for column in sheet[1])
    logger.info('Column Names: %s', column_names)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(column_names, (value for value in row))))


def get_requirements_from_wb(workbook: BytesIO, spec_sheets: list[str]) -> list[dict]:
    wb = openpyxl.load_workbook(workbook)
    sheet = wb.active

    customer = sheet.cell(row=8, column=4).value
    add_column_names = tuple(
        sheet.cell(row=2, column=sheet.max_column-i).value
        for i in reversed(range(1, 6))
    )
    sheet.delete_rows(1, 15)
    sheet.delete_rows(sheet.max_row, 1)

    column_names = (
        tuple(
            column.value for column in sheet[1] if column.value
        ) + add_column_names
    )
    logger.debug('Column Names: %s', column_names)
    logger.debug('Column Len: %s', len(column_names))

    reqs = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            spec_sheet = next(spec_sheets)
        else:
            source = dict(zip_longest(column_names, (value for value in row if value)))
            source['customer'] = customer
            source['spec_sheet'] = spec_sheet
            reqs.append(source)

    return reqs


def get_rules_from_wb(workbook_name: BytesIO) -> list[dict]:
    wb = openpyxl.load_workbook(workbook_name)
    rules = []

    for sheet in wb:
        sheet.delete_cols(0)
        sheet.delete_rows(1)

        column_names = tuple(column.value for column in sheet[1])
        logger.info('Column Names: %s', column_names)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            source = dict(zip(column_names, (value for value in row)))
            rule = {'sheet': sheet.title, **source}
            rules.append(rule)

    return rules


def get_matches_by_embedding_vector(
    index_name: str,
    embeddings: list[list[float]],
    *,
    k: Optional[int] = 1,
    score: Optional[float] = 0.0,
    vector_field_name: Optional[str] = VECTOR_FIELD_NAME
) -> list[list[dict[str, str | dict]]]:
    queries = []

    for embedding in embeddings:
        queries.append(dict())
        query = {
            "size": k,
            "min_score": score,
            "query": {
                "knn": {
                    vector_field_name: {"vector": embedding, "k": k}
                }
            }
        }
        queries.append(query)

    return get_hits_by_msearch(index_name, queries)


def _get_spec_sheets_queries(
    embeddings: list[list[float]],
    spec_sheet: str,
    *,
    k: Optional[int] = 1,
    score: Optional[float] = 0.0,
    vector_field_name: Optional[str] = 'vector_field'
) -> list[dict[str, str | dict]]:
    queries = []

    for embedding in embeddings:
        queries.append(dict())
        query = {
            "size": k,
            "min_score": score,
            "query": {
                "bool": {
                    "filter": {
                        "term": {
                            "spec_sheet.keyword": spec_sheet
                        }
                    },
                    "must": [
                        {
                            "knn": {
                                vector_field_name: {
                                    "vector": embedding,
                                    "k": 1
                                }
                            }
                        }
                    ]
                }
            }
        }
        queries.append(query)

    return queries


def get_spec_sheet_matches(
    index_name: str,
    embeddings: list[list[float]],
    spec_sheet: str,
    *,
    k: Optional[int] = 1,
    score: Optional[float] = 0.0,
    vector_field_name: Optional[str] = VECTOR_FIELD_NAME
) -> list[list[dict[str, str | dict]]]:
    queries = _get_spec_sheets_queries(
        embeddings,
        spec_sheet,
        k=k,
        score=score,
        vector_field_name=vector_field_name
    )
    logger.debug(queries)
    return get_hits_by_msearch(index_name, queries)
